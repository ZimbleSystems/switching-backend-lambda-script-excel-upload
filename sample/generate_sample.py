"""Build the full 153-row vertical onboarding workbook (Sl No 1–153).

One sheet "Onboarding": Page | Attribute | Required/Optional | Remarks (value).
Row order must match src/excel_parser.py PAGE_SEQUENCES (positional mapping).

Run:  python sample/generate_sample.py
Out:   sample/sample-vertical.xlsx  (153 data rows + header)"""
import sys
from pathlib import Path

from openpyxl import Workbook

_ROOT = Path(__file__).resolve().parent.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))
from src.excel_parser import PAGE_SEQUENCES  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUT = Path(__file__).parent / "sample-vertical.xlsx"
OUT_ALT = Path(__file__).parent / "sample-vertical-full-153.xlsx"

HEADERS = ["Sl No", "Page", "Attribute", "Required / Optional", "Remarks"]


# (Page, Attribute, Required/Optional, Remarks)  -- order MATTERS, must match
# the canonical sequence in src/excel_parser.py
ROWS = [
    # ---------------- Merchant ----------------
    ("Merchant", "Organization",            "Required", 6032),
    ("Merchant", "Merchant ID",             "Required", "01"),
    ("Merchant", "Merchant Name",           "Required", "Mas Bodega y Logistica"),
    ("Merchant", "Status",                  "Required", "Activo"),
    ("Merchant", "Governing State",         "Required", "Nuevo Leon"),
    ("Merchant", "Time Zone",               "Optional", "America/Monterrey"),
    ("Merchant", "Merchant Demographic ID", "Optional", "9"),
    ("Merchant", "Address Type",            "Optional", "Physical"),
    ("Merchant", "Address Line",            "Optional", "Av. Paseo de los Leones #2810"),
    ("Merchant", "City",                    "Optional", "Monterrey"),
    ("Merchant", "State",                   "Optional", "Nuevo Leon"),
    ("Merchant", "Postal Code",             "Optional", "64640"),
    ("Merchant", "Country",                 "Optional", "Mexico"),
    ("Merchant", "Language",                "Optional", "Espanol"),
    ("Merchant", "Longitude",               "Optional", ""),
    ("Merchant", "Longitude Direction",     "Optional", ""),
    ("Merchant", "Latitude",                "Optional", ""),
    ("Merchant", "Latitude Direction",      "Optional", ""),
    ("Merchant", "Phone Country",           "Optional", "Mexico"),
    ("Merchant", "Phone Number",            "Optional", "8112345678"),
    ("Merchant", "Type",                    "Optional", "Linea fija"),
    ("Merchant", "Email Address",           "Optional", "merchant@iconn.com.mx"),
    ("Merchant", "Type",                    "Optional", "Organizacional"),
    ("Merchant", "Social Media",            "Optional", ""),
    ("Merchant", "Logo",                    "Optional", ""),
    ("Merchant", "Photo",                   "Optional", ""),
    ("Merchant", "Facebook URL",            "Optional", ""),
    ("Merchant", "Instagram URL",           "Optional", ""),
    ("Merchant", "X (Twitter)",             "Optional", ""),
    ("Merchant", "Google ID",               "Optional", ""),
    ("Merchant", "Channel Name",            "Optional", "WEB"),
    ("Merchant", "Channel Identifier",      "Optional", "MASBODEGA-WEB"),
    ("Merchant", "Chain ID",                "Optional", "0"),
    ("Merchant", "IVA (Table)",             "Optional", ""),
    ("Merchant", "SKU (Table)",             "Optional", ""),
    ("Merchant", "Coupon (Table)",          "Optional", ""),
    ("Merchant", "Connector (Table)",       "Optional", ""),
    ("Merchant", "Merchant Criteria(Table)","Optional", "100"),

    # ---------------- Store (Sl 39–78) ----------------
    ("Store", "Store ID",                   "Required", "120"),
    ("Store", "Store Name",                 "Required", "MercaDia Villas de Alcala"),
    ("Store", "Status",                     "Required", "Activo"),
    ("Store", "Merchant",                   "Required", "01"),
    ("Store", "Governing State",            "Required", "Nuevo Leon"),
    ("Store", "Store Demographic ID",       "Required", "19"),
    ("Store", "Address Type",               "Required", "Physical"),
    ("Store", "Address Line",               "Required",
        "Calle Hibonita #300, Av. Villas de Alcala, Colonia Villas del Alcala, 2da etapa, Garcia, N.L."),
    ("Store", "City",                       "Optional", ""),
    ("Store", "State",                      "Required", "Nuevo Leon"),
    ("Store", "Postal Code",                "Optional", ""),
    ("Store", "Country",                    "Required", "Mexico"),
    ("Store", "Language",                   "Required", "Espanol"),
    ("Store", "Longitude",                  "Optional", ""),
    ("Store", "Longitude Direction",        "Optional", ""),
    ("Store", "Latitude",                   "Optional", ""),
    ("Store", "Latitude Direction",         "Optional", ""),
    ("Store", "Location Identifier Type",   "Optional", ""),
    ("Store", "Location Identifier Value",  "Optional", ""),
    ("Store", "Phone Country",              "Required", "Mexico"),
    ("Store", "Phone Number",               "Required", "8112914684"),
    ("Store", "Phone Type",                 "Required", "Linea fija"),
    ("Store", "Email Address",              "Required", "dtpmx.md020@iconn.com.mx"),
    ("Store", "Email Type",                 "Required", "Organizacional"),
    ("Store", "Social Media",               "Optional", ""),
    ("Store", "Logo",                       "Optional", ""),
    ("Store", "Photo",                      "Optional", ""),
    ("Store", "Facebook URL",               "Optional", ""),
    ("Store", "Instagram URL",              "Optional", ""),
    ("Store", "X (Twitter)",                "Optional", ""),
    ("Store", "Google ID",                  "Optional", ""),
    ("Store", "Terminal Information",       "Optional", ""),
    ("Store", "Printer",                    "Optional", ""),
    ("Store", "Chain ID",                   "Optional", "0"),
    ("Store", "IVA (Table)",                "Optional", ""),
    ("Store", "SKU (Table)",                "Optional", ""),
    ("Store", "Coupon (Table)",             "Optional", ""),
    ("Store", "Connector (Table)",          "Optional", ""),
    ("Store", "Merchant Criteria(Table)",   "Optional", "100"),
    ("Store", "Instrument Criteria(Table)", "Optional", "300"),

    # ---------------- Chain ----------------
    ("Chain", "Chain ID",                   "Required", "0"),
    ("Chain", "Chain Name",                 "Required", "Default Chain"),
    ("Chain", "Status",                     "Required", "Active"),
    ("Chain", "Governing State",            "Required", "Nuevo Leon"),
    ("Chain", "Chain Demographic ID",       "Required", "1"),
    ("Chain", "Address Type",               "Required", "Physical"),
    ("Chain", "Address Line",               "Required", "Av. Paseo de los Leones #2810"),
    ("Chain", "City",                       "Optional", "Monterrey"),
    ("Chain", "State",                      "Required", "Nuevo Leon"),
    ("Chain", "Postal Code",                "Optional", "64640"),
    ("Chain", "Country",                    "Required", "Mexico"),
    ("Chain", "Language",                   "Required", "Espanol"),
    ("Chain", "Longitude",                  "Optional", ""),
    ("Chain", "Longitude Direction",        "Optional", ""),
    ("Chain", "Latitude",                   "Optional", ""),
    ("Chain", "Latitude Direction",         "Optional", ""),
    ("Chain", "Phone Country",              "Required", "Mexico"),
    ("Chain", "Phone Number",               "Required", "8110000000"),
    ("Chain", "Phone Type",                 "Required", "Linea fija"),
    ("Chain", "Email Address",              "Required", "chain@iconn.com.mx"),
    ("Chain", "Email Type",                 "Required", "Organizacional"),
    ("Chain", "Social Media",               "Optional", ""),
    ("Chain", "Logo",                       "Optional", ""),
    ("Chain", "Photo",                      "Optional", ""),
    ("Chain", "Facebook URL",               "Optional", ""),
    ("Chain", "Instagram URL",              "Optional", ""),
    ("Chain", "X (Twitter)",                "Optional", ""),
    ("Chain", "Google ID",                  "Optional", ""),
    ("Chain", "Channel Name",               "Optional", ""),
    ("Chain", "Channel Identifier",         "Optional", ""),
    ("Chain", "Chain ID",                   "Optional", ""),  # parent chain link
    ("Chain", "IVA (Table)",                "Optional", ""),
    ("Chain", "SKU (Table)",                "Optional", ""),
    ("Chain", "Coupon (Table)",             "Optional", ""),
    ("Chain", "Connector (Table)",          "Optional", ""),
    ("Chain", "Merchant Criteria(Table)",   "Optional", "100"),
    ("Chain", "Instrument Criteria(Table)", "Optional", "300"),

    # ---------------- Merchant Criteria ----------------
    ("Merchant Criteria", "Criteria ID",            "Required", "100"),
    ("Merchant Criteria", "Criteria Description",   "Required", "Criteria 100"),
    ("Merchant Criteria", "List of countries to be excluded or included", "Optional", ""),
    ("Merchant Criteria", "List of states to be excluded or included",    "Optional", ""),
    ("Merchant Criteria", "List of currencies to be excluded or included","Optional", ""),
    ("Merchant Criteria", "List or range of MCC's to be excluded or included", "Optional", ""),
    ("Merchant Criteria", "International Type (Purchase Type)", "Optional", ""),
    ("Merchant Criteria", "Purchase Type (Purchase Type)",      "Optional", ""),
    ("Merchant Criteria", "International Type(Entry type)",     "Optional", ""),
    ("Merchant Criteria", "Entry Type(Entry type)",             "Optional", ""),
    ("Merchant Criteria", "International Type(Limit Type)",     "Optional", ""),
    ("Merchant Criteria", "Limit Type(Limit Type)",             "Optional", ""),
    ("Merchant Criteria", "Limit Type(Transaction Limit)",      "Optional", ""),
    ("Merchant Criteria", "Amount(Transaction Limit)",          "Optional", ""),
    ("Merchant Criteria", "Limit(Transaction Limit)",           "Optional", ""),
    ("Merchant Criteria", "Block CashBack?",                    "Optional", "No"),
    ("Merchant Criteria", "Block Installment",                  "Optional", "No"),
    ("Merchant Criteria", "Block International",                "Optional", "Yes"),

    # ---------------- Instrument Criteria ----------------
    ("Instrument Criteria", "Criteria ID",                              "Required", "300"),
    ("Instrument Criteria", "Criteria Description",                     "Required", "inst300"),
    ("Instrument Criteria", "Limit Type (Timed Transaction Limit)",     "Optional", "C"),
    ("Instrument Criteria", "Transaction Amount (Timed Transaction Limit)", "Optional", 5000),
    ("Instrument Criteria", "Transaction Count (Timed Transaction Limit)", "Optional", 5),
    ("Instrument Criteria", "Time Limit (Timed Transaction Limit)",     "Optional", 30),
    ("Instrument Criteria", "Time Unit (Timed Transaction Limit)",      "Optional", "Minutes"),
    ("Instrument Criteria", "Limit Type (Daily Limit)",                 "Optional", ""),
    ("Instrument Criteria", "Transaction Amount (DailyLimit)",          "Optional", ""),
    ("Instrument Criteria", "Transaction Count (Daily Limit)",          "Optional", ""),
    ("Instrument Criteria", "Channel (Suspension Control)",             "Optional", ""),
    ("Instrument Criteria", "Response Code (Suspension Control)",       "Optional", ""),
    ("Instrument Criteria", "Number of Declines (Suspension Control)",  "Optional", 3),
    ("Instrument Criteria", "Tracking Duration (Suspension Control)",   "Optional", 10),
    ("Instrument Criteria", "Tracking Time Unit (Suspension Control)",  "Optional", "Minutes"),
    ("Instrument Criteria", "Suspension Type (Suspension Control)",     "Optional", "Temporary"),
    ("Instrument Criteria", "Suspension Duration (Suspension Control)", "Optional", 60),
    ("Instrument Criteria", "Suspension Time Unit (Suspension Control)","Optional", "Minutes"),
    ("Instrument Criteria", "Check for Expiry",                         "Optional", "Yes"),
    ("Instrument Criteria", "Validate Instrument",                      "Optional", "Yes"),
]


def main() -> None:
    expected = sum(len(s) for s in PAGE_SEQUENCES.values())
    if len(ROWS) != expected:
        raise SystemExit(f"ROWS has {len(ROWS)} entries but PAGE_SEQUENCES expects {expected}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Onboarding"

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, (page, attribute, req, remarks) in enumerate(ROWS, start=2):
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=page)
        ws.cell(row=r, column=3, value=attribute)
        ws.cell(row=r, column=4, value=req)
        ws.cell(row=r, column=5, value=remarks)

    widths = {1: 7, 2: 22, 3: 50, 4: 20, 5: 60}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    target = OUT
    try:
        wb.save(target)
    except PermissionError:
        target = OUT_ALT
        wb.save(target)
        print(f"Note: close {OUT.name} in Excel if open; wrote alternate file instead.")
    print(f"Wrote {target} ({len(ROWS)} attribute rows, Sl No 1–{len(ROWS)})")


if __name__ == "__main__":
    main()
