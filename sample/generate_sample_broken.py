"""Same template but with deliberately missing/invalid required values
so we can demonstrate the validation/skip flow."""
from pathlib import Path

from openpyxl import Workbook

from generate_sample import HEADERS, ROWS  # type: ignore

OUT = Path(__file__).parent / "sample-vertical-broken.xlsx"

OVERRIDES = {
    ("Merchant", "Organization"):       "",          # required, now blank
    ("Merchant", "Status"):             "Activado",  # not in transformer map -> stays invalid
    ("Merchant Criteria", "Block CashBack?"): "Maybe",  # invalid enum
    ("Merchant Criteria", "Block Installment"): "",     # required, blank
    ("Store", "Status"):                "",          # required, now blank
    ("Chain", "Status"):                "Activated", # invalid
}


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Onboarding"

    for c, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=c, value=h)

    seen_per_attr = {}
    for r, (page, attribute, req, remarks) in enumerate(ROWS, start=2):
        key = (page, attribute)
        seen_per_attr[key] = seen_per_attr.get(key, 0) + 1
        if key in OVERRIDES:
            remarks = OVERRIDES[key]
        ws.cell(row=r, column=1, value=r - 1)
        ws.cell(row=r, column=2, value=page)
        ws.cell(row=r, column=3, value=attribute)
        ws.cell(row=r, column=4, value=req)
        ws.cell(row=r, column=5, value=remarks)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
